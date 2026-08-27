from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import tempfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import uuid4

from flask import Blueprint, jsonify, request
from flask_login import current_user, login_required
from openpyxl import load_workbook
from sqlalchemy import or_

from .access import support_access_is_active_for_current_user
from .audit import record_audit_event
from .costing import stock_unit_cost_from_purchase
from .extensions import csrf, db
from .models import (
    AuditEvent,
    DataImportChange,
    DataImportFile,
    DataImportIssue,
    DataImportJob,
    DataImportMapping,
    DataImportRow,
    InventoryItem,
    Organization,
    OrganizationMembership,
    RestaurantLocation,
    Supplier,
    SupplierItemMapping,
)
from .tenant_context import apply_org_tenant_context
from .utils import (
    get_platform_role,
    get_current_organization_bundle,
    isoformat,
    json_error,
    serialize_inventory_item,
    serialize_location,
    serialize_organization,
    serialize_supplier,
    serialize_supplier_item_mapping,
    utc_now,
)

bp = Blueprint("imports", __name__)

SUPPORTED_ENTITY_SCOPES = {
    "supplier": "Supplier",
    "inventory_item": "Inventory item",
    "supplier_item_mapping": "Supplier-item mapping",
    "opening_inventory": "Opening inventory",
}

ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
MAX_UPLOAD_BYTES = int(os.environ.get("FLOWTALLY_IMPORT_MAX_BYTES", str(10 * 1024 * 1024)))

ENTITY_FIELDS: dict[str, list[dict[str, Any]]] = {
    "supplier": [
        {"field": "name", "label": "Supplier name", "required": True},
        {"field": "categoryFocus", "label": "Category focus", "required": False},
        {"field": "contactName", "label": "Contact name", "required": False},
        {"field": "contactPhone", "label": "Contact phone", "required": False},
        {"field": "contactEmail", "label": "Contact email", "required": False},
        {"field": "orderingNotes", "label": "Ordering notes", "required": False},
        {"field": "notes", "label": "Notes", "required": False},
    ],
    "inventory_item": [
        {"field": "name", "label": "Item name", "required": True},
        {"field": "locationName", "label": "Location name", "required": True},
        {"field": "supplierName", "label": "Supplier name", "required": False},
        {"field": "category", "label": "Category", "required": False},
        {"field": "stockUnit", "label": "Stock unit", "required": False},
        {"field": "currentOnHand", "label": "Current on hand", "required": False},
        {"field": "minQuantity", "label": "Minimum quantity", "required": False},
        {"field": "parLevel", "label": "PAR level", "required": False},
        {"field": "latestPurchasePrice", "label": "Latest purchase price", "required": False},
    ],
    "supplier_item_mapping": [
        {"field": "supplierName", "label": "Supplier name", "required": True},
        {"field": "inventoryItemName", "label": "Inventory item name", "required": True},
        {"field": "supplierItemName", "label": "Supplier item name", "required": True},
        {"field": "purchaseUnit", "label": "Purchase unit", "required": False},
        {"field": "inventoryUnit", "label": "Inventory unit", "required": False},
        {"field": "conversionFactor", "label": "Conversion factor", "required": False},
    ],
    "opening_inventory": [
        {"field": "itemName", "label": "Item name", "required": True},
        {"field": "locationName", "label": "Location name", "required": True},
        {"field": "currentOnHand", "label": "Opening quantity", "required": True},
        {"field": "stockUnit", "label": "Stock unit", "required": False},
    ],
}

ALLOWED_UNITS = {"each", "case", "box", "bottle", "bag", "kg", "g", "lb", "oz", "liter", "litre", "ml", "dozen", "pack", "sheet"}


def _normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _normalize_key(value: Any) -> str:
    return _normalize_text(value).lower()


def _column_key(value: Any) -> str:
    return _normalize_key(value).replace(" ", "_")


def _to_decimal(value: Any, *, field_name: str, allow_negative: bool = False) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name} must be a valid number.") from exc
    if not allow_negative and decimal_value < 0:
        raise ValueError(f"{field_name} cannot be negative.")
    return decimal_value


def _max_upload_bytes() -> int:
    return MAX_UPLOAD_BYTES


def _import_root() -> Path:
    root = Path(tempfile.gettempdir()) / "flowtally-imports"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _tenant_temp_dir(organization_id: int) -> Path:
    path = _import_root() / f"org-{organization_id}"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _build_storage_path(organization_id: int, suffix: str) -> Path:
    return _tenant_temp_dir(organization_id) / f"job-{uuid4().hex}{suffix}"


def _sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _read_upload_bytes(uploaded_file) -> bytes:
    content = uploaded_file.read()
    if not content:
        raise ValueError("The uploaded file is empty.")
    if len(content) > _max_upload_bytes():
        raise ValueError(f"Uploaded file is too large. Maximum size is {_max_upload_bytes() // (1024 * 1024)} MB.")
    return content


def _validate_extension(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("Only CSV and XLSX files are supported.")
    return suffix


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("The uploaded CSV file must include a header row.")
    columns = [_normalize_text(column) for column in reader.fieldnames if column is not None]
    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        rows.append({str(key): value for key, value in raw_row.items() if key is not None})
    return columns, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    try:
        header = next(iterator)
    except StopIteration as exc:
        raise ValueError("The uploaded XLSX file is empty.") from exc
    if not header:
        raise ValueError("The uploaded XLSX file must include a header row.")
    columns = [_normalize_text(column) for column in header if column is not None]
    rows: list[dict[str, Any]] = []
    for row in iterator:
        row_payload: dict[str, Any] = {}
        for index, column_name in enumerate(header):
            if column_name is None:
                continue
            row_payload[str(column_name)] = row[index] if index < len(row) else None
        rows.append(row_payload)
    return columns, rows


def _parse_spreadsheet(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]], str]:
    suffix = _validate_extension(filename)
    if suffix == ".csv":
        columns, rows = _parse_csv(content)
        return columns, rows, "csv"
    columns, rows = _parse_xlsx(content)
    return columns, rows, "xlsx"


def _row_source_value(row: dict[str, Any], field_name: str) -> Any:
    if field_name in row:
        return row[field_name]
    normalized_target = _column_key(field_name)
    for key, value in row.items():
        if _column_key(key) == normalized_target:
            return value
    return None


def _find_organization_for_request(organization_id: int | None) -> Organization | None:
    if organization_id is not None:
        return Organization.query.filter_by(id=organization_id).first()
    organization, _, _ = get_current_organization_bundle()
    return organization


def _current_membership_for_org(organization_id: int) -> OrganizationMembership | None:
    return OrganizationMembership.query.filter_by(organization_id=organization_id, user_id=current_user.id).first()


def _can_access_imports(organization: Organization | None) -> tuple[bool, str]:
    if organization is None:
        return False, "Organization not found."
    membership = _current_membership_for_org(organization.id)
    if membership and membership.role == "owner":
        return True, ""
    role = get_platform_role(current_user.id)
    if role and role.is_active and role.role == "setup_admin":
        return True, ""
    if role and role.is_active and role.role == "support" and support_access_is_active_for_current_user():
        active_org, _, _ = get_current_organization_bundle()
        if active_org and active_org.id == organization.id:
            return True, ""
    return False, "You do not have permission to manage imports for this organization."


def _require_import_access(organization_id: int | None) -> tuple[Organization | None, tuple[dict[str, Any], int] | None]:
    organization = _find_organization_for_request(organization_id)
    allowed, message = _can_access_imports(organization)
    if not allowed:
        return organization, json_error(message or "Access denied.", 403)
    apply_org_tenant_context(organization, access_scope="setup" if get_platform_role(current_user.id) and get_platform_role(current_user.id).role == "setup_admin" else None)
    return organization, None


def _safe_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _mapping_config_from_job(job: DataImportJob) -> dict[str, Any]:
    return dict(job.mapping_json or {})


def _field_mappings_from_config(config: dict[str, Any]) -> dict[str, str]:
    mappings = config.get("fieldMappings") or {}
    return {str(key): str(value) for key, value in mappings.items() if str(value).strip()}


def _fixed_values_from_config(config: dict[str, Any]) -> dict[str, Any]:
    values = config.get("fixedValues") or {}
    return {str(key): value for key, value in values.items()}


def _entity_fields(entity_scope: str) -> list[dict[str, Any]]:
    return list(ENTITY_FIELDS.get(entity_scope, ENTITY_FIELDS["supplier"]))


def _field_spec(entity_scope: str, field_name: str) -> dict[str, Any] | None:
    return next((field for field in _entity_fields(entity_scope) if field["field"] == field_name), None)


def _column_list(rows: list[dict[str, Any]]) -> list[str]:
    seen: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.append(key)
    return seen


def _row_fingerprint(entity_scope: str, normalized_row: dict[str, Any]) -> str:
    payload = json.dumps({"entityScope": entity_scope, "row": normalized_row}, sort_keys=True, default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _job_to_summary(job: DataImportJob) -> dict[str, Any]:
    columns = list((job.summary_json or {}).get("sourceColumns") or [])
    sample_rows = list((job.summary_json or {}).get("sampleRows") or [])
    return {
        "id": job.id,
        "organizationId": job.organization_id,
        "createdByUserId": job.created_by_user_id,
        "approvedByUserId": job.approved_by_user_id,
        "sourceType": job.source_type,
        "sourceFileName": job.source_file_name,
        "sourceFileExtension": job.source_file_extension,
        "sourceMimeType": job.source_mime_type,
        "sourceHash": job.source_hash,
        "storagePath": job.storage_path,
        "status": job.status,
        "entityScope": job.entity_scope,
        "mapping": dict(job.mapping_json or {}),
        "summary": dict(job.summary_json or {}),
        "rowCount": job.row_count,
        "previewRowCount": job.preview_row_count,
        "appliedRowCount": job.applied_row_count,
        "blockedRowCount": job.blocked_row_count,
        "warningCount": job.warning_count,
        "approvedAt": isoformat(job.approved_at),
        "executedAt": isoformat(job.executed_at),
        "rolledBackAt": isoformat(job.rolled_back_at),
        "batchId": job.batch_id,
        "rollbackBlockers": list(job.rollback_blockers_json or []),
        "sourceColumns": columns,
        "sampleRows": sample_rows,
        "createdAt": isoformat(job.created_at),
        "updatedAt": isoformat(job.updated_at),
    }


def _serialize_issue(issue: DataImportIssue) -> dict[str, Any]:
    return {
        "id": issue.id,
        "dataImportRowId": issue.data_import_row_id,
        "severity": issue.severity,
        "fieldName": issue.field_name,
        "code": issue.code,
        "message": issue.message,
        "createdAt": isoformat(issue.created_at),
        "updatedAt": isoformat(issue.updated_at),
    }


def _serialize_row(row: DataImportRow) -> dict[str, Any]:
    return {
        "id": row.id,
        "dataImportJobId": row.data_import_job_id,
        "rowNumber": row.row_number,
        "entityType": row.entity_type,
        "sourceRow": dict(row.source_row_json or {}),
        "normalizedRow": dict(row.normalized_row_json or {}),
        "rowFingerprint": row.row_fingerprint,
        "status": row.status,
        "targetEntityType": row.target_entity_type,
        "targetEntityId": row.target_entity_id,
        "issueSummary": row.issue_summary,
        "warningCount": row.warning_count,
        "blockedCount": row.blocked_count,
        "canRollback": bool(row.can_rollback),
        "issues": [_serialize_issue(issue) for issue in row.issues],
        "changes": [_serialize_change(change) for change in row.changes],
        "createdAt": isoformat(row.created_at),
        "updatedAt": isoformat(row.updated_at),
    }


def _serialize_change(change: DataImportChange) -> dict[str, Any]:
    return {
        "id": change.id,
        "dataImportJobId": change.data_import_job_id,
        "dataImportRowId": change.data_import_row_id,
        "entityType": change.entity_type,
        "changeType": change.change_type,
        "targetEntityId": change.target_entity_id,
        "rowFingerprint": change.row_fingerprint,
        "previous": dict(change.previous_json or {}),
        "applied": dict(change.applied_json or {}),
        "rollbackable": bool(change.rollbackable),
        "status": change.status,
        "createdAt": isoformat(change.created_at),
        "updatedAt": isoformat(change.updated_at),
    }


def _serialize_job_detail(job: DataImportJob) -> dict[str, Any]:
    return {
        **_job_to_summary(job),
        "organization": serialize_organization(job.organization) if job.organization else None,
        "files": [
            {
                "id": file.id,
                "role": file.role,
                "originalFileName": file.original_file_name,
                "storagePath": file.storage_path,
                "sha256": file.sha256,
                "byteSize": file.byte_size,
                "mimeType": file.mime_type,
                "uploadedAt": isoformat(file.uploaded_at),
                "createdAt": isoformat(file.created_at),
                "updatedAt": isoformat(file.updated_at),
            }
            for file in job.files
        ],
        "rows": [_serialize_row(row) for row in job.rows],
        "changes": [_serialize_change(change) for change in job.changes],
        "mappings": [
            {
                "id": mapping.id,
                "sourceColumnName": mapping.source_column_name,
                "targetFieldName": mapping.target_field_name,
                "mappingType": mapping.mapping_type,
                "fixedValue": dict(mapping.fixed_value_json or {}),
                "displayOrder": mapping.display_order,
                "isRequired": bool(mapping.is_required),
                "createdAt": isoformat(mapping.created_at),
                "updatedAt": isoformat(mapping.updated_at),
            }
            for mapping in job.mappings
        ],
        "auditEvents": [
            {
                "id": event.id,
                "eventType": event.event_type,
                "entityType": event.entity_type,
                "entityId": event.entity_id,
                "metadata": dict(event.metadata_json or {}),
                "createdAt": isoformat(event.created_at),
            }
            for event in AuditEvent.query.filter_by(organization_id=job.organization_id)
            .order_by(AuditEvent.created_at.desc(), AuditEvent.id.desc())
            .limit(20)
            .all()
        ],
    }


def _load_job(job_id: int) -> DataImportJob | None:
    return DataImportJob.query.filter_by(id=job_id).first()


def _job_or_404(job_id: int) -> tuple[DataImportJob | None, tuple[dict[str, Any], int] | None]:
    job = _load_job(job_id)
    if job is None:
        return None, json_error("Import job not found.", 404)
    organization, error = _require_import_access(job.organization_id)
    if error:
        return job, error
    return job, None


def _apply_mapping(job: DataImportJob, mapping_payload: dict[str, Any]) -> None:
    entity_scope = str(mapping_payload.get("entityScope") or job.entity_scope).strip() or job.entity_scope
    if entity_scope not in SUPPORTED_ENTITY_SCOPES:
        raise ValueError("Unknown entity scope.")
    field_mappings = mapping_payload.get("fieldMappings") or {}
    fixed_values = mapping_payload.get("fixedValues") or {}
    if not isinstance(field_mappings, dict) or not isinstance(fixed_values, dict):
        raise ValueError("Mapping payload must include fieldMappings and fixedValues objects.")

    allowed_fields = {field["field"] for field in _entity_fields(entity_scope)}
    mapping_rows: list[DataImportMapping] = []
    DataImportMapping.query.filter_by(data_import_job_id=job.id).delete(synchronize_session=False)
    display_order = 1
    for target_field, source_column in field_mappings.items():
        target_field_name = str(target_field).strip()
        if target_field_name not in allowed_fields:
            raise ValueError(f"Unsupported target field: {target_field_name}.")
        field_spec = _field_spec(entity_scope, target_field_name)
        mapping_rows.append(
            DataImportMapping(
                job=job,
                source_column_name=str(source_column).strip(),
                target_field_name=target_field_name,
                mapping_type="manual",
                fixed_value_json={},
                display_order=display_order,
                is_required=bool(field_spec and field_spec.get("required")),
            )
        )
        display_order += 1
    for fixed_field, fixed_value in fixed_values.items():
        fixed_field_name = str(fixed_field).strip()
        if fixed_field_name not in allowed_fields:
            raise ValueError(f"Unsupported fixed field: {fixed_field_name}.")
        field_spec = _field_spec(entity_scope, fixed_field_name)
        mapping_rows.append(
            DataImportMapping(
                job=job,
                source_column_name="",
                target_field_name=fixed_field_name,
                mapping_type="fixed",
                fixed_value_json={"value": fixed_value},
                display_order=display_order,
                is_required=bool(field_spec and field_spec.get("required")),
            )
        )
        display_order += 1

    job.entity_scope = entity_scope
    job.mapping_json = {
        "entityScope": entity_scope,
        "fieldMappings": {str(key): str(value) for key, value in field_mappings.items()},
        "fixedValues": {str(key): value for key, value in fixed_values.items()},
    }
    job.status = "VALIDATION_REQUIRED"
    db.session.add_all(mapping_rows)


def _source_row_value(row: dict[str, Any], source_column_name: str) -> Any:
    if not source_column_name:
        return None
    if source_column_name in row:
        return row[source_column_name]
    normalized = _column_key(source_column_name)
    for key, value in row.items():
        if _column_key(key) == normalized:
            return value
    return None


def _lookup_location_by_name(organization_id: int, location_name: str) -> RestaurantLocation | None:
    normalized = _normalize_key(location_name)
    if not normalized:
        return None
    return (
        RestaurantLocation.query.filter(RestaurantLocation.organization_id == organization_id)
        .filter(db.func.lower(db.func.trim(RestaurantLocation.name)) == normalized)
        .first()
    )


def _lookup_supplier_by_name(organization_id: int, supplier_name: str) -> Supplier | None:
    normalized = _normalize_key(supplier_name)
    if not normalized:
        return None
    return (
        Supplier.query.filter(Supplier.organization_id == organization_id)
        .filter(db.func.lower(db.func.trim(Supplier.normalized_name)) == normalized)
        .first()
    )


def _lookup_item_by_name(organization_id: int, location_id: int, item_name: str) -> InventoryItem | None:
    normalized = _normalize_key(item_name)
    if not normalized:
        return None
    return (
        InventoryItem.query.filter(
            InventoryItem.organization_id == organization_id,
            InventoryItem.location_id == location_id,
            db.func.lower(db.func.trim(InventoryItem.normalized_name)) == normalized,
        )
        .first()
    )


def _lookup_mapping(organization_id: int, supplier_id: int, inventory_item_id: int, supplier_item_name: str) -> SupplierItemMapping | None:
    normalized = _normalize_key(supplier_item_name)
    if not normalized:
        return None
    return (
        SupplierItemMapping.query.filter(
            SupplierItemMapping.organization_id == organization_id,
            SupplierItemMapping.supplier_id == supplier_id,
            SupplierItemMapping.inventory_item_id == inventory_item_id,
            db.func.lower(db.func.trim(SupplierItemMapping.normalized_supplier_item_name)) == normalized,
        )
        .first()
    )


def _serialize_entity(entity: Any, entity_type: str) -> dict[str, Any]:
    if entity_type == "supplier":
        return serialize_supplier(entity)
    if entity_type == "inventory_item":
        return serialize_inventory_item(entity)
    if entity_type == "supplier_item_mapping":
        return serialize_supplier_item_mapping(entity)
    if entity_type == "opening_inventory":
        return serialize_inventory_item(entity)
    return {}


def _save_upload(job: DataImportJob, filename: str, content: bytes) -> DataImportFile:
    suffix = _validate_extension(filename)
    storage_path = _build_storage_path(job.organization_id, suffix)
    storage_path.write_bytes(content)
    uploaded_file = DataImportFile(
        job=job,
        role="source",
        original_file_name=Path(filename).name,
        storage_path=str(storage_path),
        sha256=_sha256(content),
        byte_size=len(content),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if suffix == ".xlsx" else "text/csv",
    )
    db.session.add(uploaded_file)
    return uploaded_file


def _job_file_content(job: DataImportJob) -> bytes:
    file = next((entry for entry in job.files if entry.role == "source"), None)
    if file is None:
        raise ValueError("Source file is missing.")
    return Path(file.storage_path).read_bytes()


def _build_normalized_row(job: DataImportJob, source_row: dict[str, Any]) -> dict[str, Any]:
    config = _mapping_config_from_job(job)
    field_mappings = _field_mappings_from_config(config)
    fixed_values = _fixed_values_from_config(config)
    normalized: dict[str, Any] = {}
    for field_spec in _entity_fields(job.entity_scope):
        field_name = field_spec["field"]
        if field_name in fixed_values:
            normalized[field_name] = fixed_values[field_name]
            continue
        source_column = field_mappings.get(field_name)
        normalized[field_name] = _source_row_value(source_row, source_column) if source_column else None
    return normalized


def _row_issues_for_entity(job: DataImportJob, organization: Organization, row: dict[str, Any], normalized: dict[str, Any]) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    entity = job.entity_scope

    def add_issue(field_name: str, code: str, message: str, *, severity: str = "blocked") -> None:
        issues.append({"fieldName": field_name, "code": code, "message": message, "severity": severity})

    if entity == "supplier":
        if not _normalize_text(normalized.get("name")):
            add_issue("name", "required", "Supplier name is required.")
        if normalized.get("contactEmail") and "@" not in str(normalized.get("contactEmail")):
            add_issue("contactEmail", "invalid_email", "Enter a valid contact email.", severity="warning")
        return issues

    if entity in {"inventory_item", "opening_inventory"}:
        name_field = "name" if entity == "inventory_item" else "itemName"
        location_value = normalized.get("locationName")
        if not _normalize_text(normalized.get(name_field)):
            add_issue(name_field, "required", "Item name is required.")
        if not _normalize_text(location_value):
            add_issue("locationName", "required", "Location name is required.")
        location = _lookup_location_by_name(organization.id, str(location_value or ""))
        if location_value and location is None:
            add_issue("locationName", "missing_location", "The selected location does not belong to this organization.")
        if normalized.get("supplierName") and _lookup_supplier_by_name(organization.id, str(normalized.get("supplierName") or "")) is None:
            add_issue("supplierName", "missing_supplier", "The selected supplier does not belong to this organization.")
        if normalized.get("stockUnit") and _normalize_key(normalized.get("stockUnit")) not in { _normalize_key(unit) for unit in ALLOWED_UNITS }:
            add_issue("stockUnit", "unknown_unit", "That stock unit is not recognized.", severity="warning")
        for numeric_field in ["currentOnHand", "minQuantity", "parLevel", "latestPurchasePrice"]:
            if numeric_field not in normalized:
                continue
            if normalized.get(numeric_field) in (None, ""):
                continue
            try:
                parsed = _to_decimal(normalized.get(numeric_field), field_name=numeric_field)
                if parsed is not None and parsed < 0:
                    add_issue(numeric_field, "negative_value", f"{numeric_field} cannot be negative.")
            except ValueError as exc:
                add_issue(numeric_field, "invalid_number", str(exc))
        return issues

    if entity == "supplier_item_mapping":
        if not _normalize_text(normalized.get("supplierName")):
            add_issue("supplierName", "required", "Supplier name is required.")
        elif _lookup_supplier_by_name(organization.id, str(normalized.get("supplierName") or "")) is None:
            add_issue("supplierName", "missing_supplier", "The selected supplier does not belong to this organization.")
        if not _normalize_text(normalized.get("inventoryItemName")):
            add_issue("inventoryItemName", "required", "Inventory item name is required.")
        else:
            location_name = str(normalized.get("locationName") or "")
            location = _lookup_location_by_name(organization.id, location_name) if location_name else None
            if location is None:
                add_issue("locationName", "missing_location", "The selected location does not belong to this organization.")
            elif _lookup_item_by_name(organization.id, location.id, str(normalized.get("inventoryItemName") or "")) is None:
                add_issue("inventoryItemName", "missing_item", "The selected inventory item does not belong to this organization.")
        if not _normalize_text(normalized.get("supplierItemName")):
            add_issue("supplierItemName", "required", "Supplier item name is required.")
        if normalized.get("conversionFactor") not in (None, ""):
            try:
                parsed = _to_decimal(normalized.get("conversionFactor"), field_name="conversionFactor")
                if parsed is not None and parsed <= 0:
                    add_issue("conversionFactor", "invalid_number", "Conversion factor must be greater than zero.")
            except ValueError as exc:
                add_issue("conversionFactor", "invalid_number", str(exc))
        return issues

    add_issue("entityScope", "unknown_entity", "That import type is not supported.")
    return issues


def _clear_preview_state(job: DataImportJob) -> None:
    DataImportIssue.query.filter(DataImportIssue.data_import_row_id.in_([row.id for row in job.rows])).delete(synchronize_session=False)
    DataImportChange.query.filter_by(data_import_job_id=job.id).delete()


def _build_preview_row(job: DataImportJob, organization: Organization, row: DataImportRow) -> None:
    normalized = dict(row.normalized_row_json or {})
    issues = _row_issues_for_entity(job, organization, dict(row.source_row_json or {}), normalized)
    row.issue_summary = "; ".join(issue["message"] for issue in issues)
    row.warning_count = sum(1 for issue in issues if issue["severity"] == "warning")
    row.blocked_count = sum(1 for issue in issues if issue["severity"] == "blocked")
    row.status = "blocked" if row.blocked_count else "ready"
    row.can_rollback = True
    for issue in issues:
        db.session.add(
            DataImportIssue(
                row=row,
                severity=issue["severity"],
                field_name=issue["fieldName"],
                code=issue["code"],
                message=issue["message"],
            )
        )

    change_type = "skip"
    target_entity_id = ""
    previous_json: dict[str, Any] = {}
    applied_json: dict[str, Any] = dict(normalized)
    rollbackable = True
    target_entity_type = job.entity_scope

    if row.blocked_count == 0:
        if job.entity_scope == "supplier":
            supplier = _lookup_supplier_by_name(organization.id, str(normalized.get("name") or ""))
            if supplier is None:
                change_type = "create"
            else:
                change_type = "update"
                target_entity_id = str(supplier.id)
                previous_json = _serialize_entity(supplier, "supplier")
        elif job.entity_scope in {"inventory_item", "opening_inventory"}:
            location = _lookup_location_by_name(organization.id, str(normalized.get("locationName") or ""))
            if location is not None:
                item_name = str(normalized.get("name") or normalized.get("itemName") or "")
                item = _lookup_item_by_name(organization.id, location.id, item_name)
                if item is None:
                    change_type = "create"
                else:
                    change_type = "update"
                    target_entity_id = str(item.id)
                    previous_json = _serialize_entity(item, "inventory_item")
            else:
                change_type = "skip"
        elif job.entity_scope == "supplier_item_mapping":
            supplier = _lookup_supplier_by_name(organization.id, str(normalized.get("supplierName") or ""))
            location_name = str(normalized.get("locationName") or "")
            inventory_item_name = str(normalized.get("inventoryItemName") or "")
            item = None
            if location_name:
                location = _lookup_location_by_name(organization.id, location_name)
                if location is not None:
                    item = _lookup_item_by_name(organization.id, location.id, inventory_item_name)
            if supplier and item:
                mapping = _lookup_mapping(organization.id, supplier.id, item.id, str(normalized.get("supplierItemName") or ""))
                if mapping is None:
                    change_type = "create"
                else:
                    change_type = "update"
                    target_entity_id = str(mapping.id)
                    previous_json = _serialize_entity(mapping, "supplier_item_mapping")
            else:
                change_type = "skip"
        if change_type == "skip":
            row.status = "blocked" if row.blocked_count else "ready"
        else:
            row.status = "ready"

    row.target_entity_type = target_entity_type
    row.target_entity_id = target_entity_id
    row.normalized_row_json = normalized
    fingerprint = _row_fingerprint(job.entity_scope, normalized)
    row.row_fingerprint = fingerprint
    db.session.add(
        DataImportChange(
            job=job,
            row=row,
            entity_type=job.entity_scope,
            change_type=change_type,
            target_entity_id=target_entity_id,
            row_fingerprint=fingerprint,
            previous_json=previous_json,
            applied_json=applied_json,
            rollbackable=rollbackable,
            status="preview",
        )
    )


def _refresh_summary(job: DataImportJob) -> None:
    rows = list(job.rows)
    job.preview_row_count = sum(1 for row in rows if row.status == "ready")
    job.blocked_row_count = sum(1 for row in rows if row.blocked_count)
    job.warning_count = sum(row.warning_count for row in rows)
    job.applied_row_count = sum(1 for row in rows if row.status == "completed")
    job.summary_json = {
        "sourceColumns": list((job.summary_json or {}).get("sourceColumns") or []),
        "sampleRows": list((job.summary_json or {}).get("sampleRows") or []),
        "rowCount": job.row_count,
        "previewRowCount": job.preview_row_count,
        "blockedRowCount": job.blocked_row_count,
        "warningCount": job.warning_count,
        "supportedEntityScopes": list(SUPPORTED_ENTITY_SCOPES.keys()),
    }


def _cleanup_source_file(job: DataImportJob) -> None:
    for file in job.files:
        if file.role != "source":
            continue
        if not file.storage_path:
            continue
        path = Path(file.storage_path)
        if path.exists() and path.is_file():
            path.unlink()
        file.storage_path = ""


def _restore_supplier(payload: dict[str, Any], supplier: Supplier) -> None:
    supplier.name = str(payload.get("name") or supplier.name)
    supplier.normalized_name = _normalize_key(payload.get("normalizedName") or payload.get("name") or supplier.name)
    supplier.category_focus = str(payload.get("categoryFocus") or "")
    supplier.contact_name = str(payload.get("contactName") or "")
    supplier.contact_phone = str(payload.get("contactPhone") or "")
    supplier.contact_email = str(payload.get("contactEmail") or "")
    supplier.ordering_notes = str(payload.get("orderingNotes") or "")
    supplier.notes = str(payload.get("notes") or "")
    supplier.is_active = bool(payload.get("isActive", True))


def _restore_inventory_item(payload: dict[str, Any], item: InventoryItem) -> None:
    item.name = str(payload.get("name") or item.name)
    item.normalized_name = _normalize_key(payload.get("normalizedName") or payload.get("name") or item.name)
    item.category = str(payload.get("category") or "Other")
    item.stock_unit = str(payload.get("stockUnit") or "each")
    item.current_on_hand = Decimal(str(payload.get("currentOnHand") or 0))
    item.min_quantity = Decimal(str(payload.get("minQuantity") or 0))
    item.par_level = Decimal(str(payload.get("parLevel") or 0))
    item.preferred_supplier_name = str(payload.get("preferredSupplierName") or "")
    item.latest_purchase_price = Decimal(str(payload.get("latestPurchasePrice") or 0))
    item.last_purchase_unit = str(payload.get("lastPurchaseUnit") or "each")
    item.last_purchase_conversion_factor = Decimal(str(payload.get("lastPurchaseConversionFactor") or 1))
    if payload.get("averageUnitCost") not in (None, ""):
        item.average_unit_cost = Decimal(str(payload.get("averageUnitCost") or 0))
    else:
        item.average_unit_cost = stock_unit_cost_from_purchase(item.latest_purchase_price, item.last_purchase_conversion_factor)
    item.active = bool(payload.get("active", True))
    item.notes = str(payload.get("notes") or "")
    item.supplier_id = payload.get("supplierId")
    item.location_id = int(payload.get("locationId") or item.location_id)


def _restore_mapping(payload: dict[str, Any], mapping: SupplierItemMapping) -> None:
    mapping.supplier_item_name = str(payload.get("supplierItemName") or mapping.supplier_item_name)
    mapping.normalized_supplier_item_name = _normalize_key(payload.get("normalizedSupplierItemName") or payload.get("supplierItemName") or mapping.supplier_item_name)
    mapping.purchase_unit = str(payload.get("purchaseUnit") or "each")
    mapping.inventory_unit = str(payload.get("inventoryUnit") or "each")
    mapping.conversion_factor = Decimal(str(payload.get("conversionFactor") or 1))


def _preview_job(job: DataImportJob, organization: Organization) -> None:
    _clear_preview_state(job)
    for row in job.rows:
        row.normalized_row_json = _build_normalized_row(job, dict(row.source_row_json or {}))
        _build_preview_row(job, organization, row)
    job.status = "READY_FOR_PREVIEW" if job.blocked_row_count == 0 else "VALIDATION_REQUIRED"
    _refresh_summary(job)


def _apply_row(job: DataImportJob, organization: Organization, row: DataImportRow, user_id: int) -> tuple[str, Any, dict[str, Any], dict[str, Any], bool]:
    normalized = dict(row.normalized_row_json or {})
    previous_json: dict[str, Any] = {}
    applied_json: dict[str, Any] = dict(normalized)
    rollbackable = True

    if job.entity_scope == "supplier":
        name = _normalize_text(normalized.get("name"))
        supplier = _lookup_supplier_by_name(organization.id, name)
        if supplier is None:
            supplier = Supplier(
                organization_id=organization.id,
                name=name,
                normalized_name=_normalize_key(name),
            )
            db.session.add(supplier)
            db.session.flush()
            applied_json = serialize_supplier(supplier)
            return "create", supplier, previous_json, applied_json, rollbackable
        previous_json = serialize_supplier(supplier)
        supplier.name = name
        supplier.normalized_name = _normalize_key(name)
        supplier.category_focus = str(normalized.get("categoryFocus") or supplier.category_focus or "Other")
        supplier.contact_name = str(normalized.get("contactName") or "")
        supplier.contact_phone = str(normalized.get("contactPhone") or "")
        supplier.contact_email = str(normalized.get("contactEmail") or "")
        supplier.ordering_notes = str(normalized.get("orderingNotes") or "")
        supplier.notes = str(normalized.get("notes") or "")
        applied_json = serialize_supplier(supplier)
        return "update", supplier, previous_json, applied_json, rollbackable

    if job.entity_scope in {"inventory_item", "opening_inventory"}:
        location = _lookup_location_by_name(organization.id, str(normalized.get("locationName") or ""))
        if location is None:
            raise ValueError("The selected location does not belong to this organization.")
        item_name = str(normalized.get("name") or normalized.get("itemName") or "").strip()
        item = _lookup_item_by_name(organization.id, location.id, item_name)
        if item is None:
            item = InventoryItem(
                organization_id=organization.id,
                location_id=location.id,
                name=item_name,
                normalized_name=_normalize_key(item_name),
            )
            db.session.add(item)
            db.session.flush()
            if normalized.get("supplierName"):
                supplier = _lookup_supplier_by_name(organization.id, str(normalized.get("supplierName") or ""))
                if supplier:
                    item.supplier_id = supplier.id
            item.category = str(normalized.get("category") or "Other")
            item.stock_unit = str(normalized.get("stockUnit") or "each")
            current_on_hand = _to_decimal(normalized.get("currentOnHand"), field_name="currentOnHand") if normalized.get("currentOnHand") not in (None, "") else None
            if current_on_hand is not None:
                item.current_on_hand = current_on_hand
            min_quantity = _to_decimal(normalized.get("minQuantity"), field_name="minQuantity") if normalized.get("minQuantity") not in (None, "") else None
            if min_quantity is not None:
                item.min_quantity = min_quantity
            par_level = _to_decimal(normalized.get("parLevel"), field_name="parLevel") if normalized.get("parLevel") not in (None, "") else None
            if par_level is not None:
                item.par_level = par_level
            latest_purchase_price = _to_decimal(normalized.get("latestPurchasePrice"), field_name="latestPurchasePrice") if normalized.get("latestPurchasePrice") not in (None, "") else None
            if latest_purchase_price is not None:
                item.latest_purchase_price = latest_purchase_price
            item.average_unit_cost = stock_unit_cost_from_purchase(item.latest_purchase_price, item.last_purchase_conversion_factor)
            applied_json = serialize_inventory_item(item)
            return "create", item, previous_json, applied_json, rollbackable
        previous_json = serialize_inventory_item(item)
        item.name = item_name
        item.normalized_name = _normalize_key(item_name)
        if normalized.get("supplierName"):
            supplier = _lookup_supplier_by_name(organization.id, str(normalized.get("supplierName") or ""))
            if supplier:
                item.supplier_id = supplier.id
        item.category = str(normalized.get("category") or item.category or "Other")
        item.stock_unit = str(normalized.get("stockUnit") or item.stock_unit or "each")
        for field_name, attr_name in [("currentOnHand", "current_on_hand"), ("minQuantity", "min_quantity"), ("parLevel", "par_level"), ("latestPurchasePrice", "latest_purchase_price")]:
            if normalized.get(field_name) not in (None, ""):
                setattr(item, attr_name, _to_decimal(normalized.get(field_name), field_name=field_name))
        if normalized.get("averageUnitCost") not in (None, ""):
            item.average_unit_cost = _to_decimal(normalized.get("averageUnitCost"), field_name="averageUnitCost")
        else:
            item.average_unit_cost = stock_unit_cost_from_purchase(item.latest_purchase_price, item.last_purchase_conversion_factor)
        applied_json = serialize_inventory_item(item)
        return "update", item, previous_json, applied_json, rollbackable

    if job.entity_scope == "supplier_item_mapping":
        supplier = _lookup_supplier_by_name(organization.id, str(normalized.get("supplierName") or ""))
        location_name = str(normalized.get("locationName") or "")
        location = _lookup_location_by_name(organization.id, location_name) if location_name else None
        item = _lookup_item_by_name(organization.id, location.id if location else 0, str(normalized.get("inventoryItemName") or "")) if location else None
        if supplier is None or item is None:
            raise ValueError("Referenced supplier or inventory item could not be found.")
        supplier_item_name = str(normalized.get("supplierItemName") or "").strip()
        mapping = _lookup_mapping(organization.id, supplier.id, item.id, supplier_item_name)
        if mapping is None:
            mapping = SupplierItemMapping(
                organization_id=organization.id,
                supplier_id=supplier.id,
                inventory_item_id=item.id,
                supplier_item_name=supplier_item_name,
                normalized_supplier_item_name=_normalize_key(supplier_item_name),
            )
            db.session.add(mapping)
            db.session.flush()
            mapping.purchase_unit = str(normalized.get("purchaseUnit") or "each")
            mapping.inventory_unit = str(normalized.get("inventoryUnit") or item.stock_unit or "each")
            if normalized.get("conversionFactor") not in (None, ""):
                mapping.conversion_factor = _to_decimal(normalized.get("conversionFactor"), field_name="conversionFactor")
            applied_json = serialize_supplier_item_mapping(mapping)
            return "create", mapping, previous_json, applied_json, rollbackable
        previous_json = serialize_supplier_item_mapping(mapping)
        mapping.supplier_item_name = supplier_item_name
        mapping.normalized_supplier_item_name = _normalize_key(supplier_item_name)
        mapping.purchase_unit = str(normalized.get("purchaseUnit") or mapping.purchase_unit or "each")
        mapping.inventory_unit = str(normalized.get("inventoryUnit") or mapping.inventory_unit or item.stock_unit or "each")
        if normalized.get("conversionFactor") not in (None, ""):
            mapping.conversion_factor = _to_decimal(normalized.get("conversionFactor"), field_name="conversionFactor")
        applied_json = serialize_supplier_item_mapping(mapping)
        return "update", mapping, previous_json, applied_json, rollbackable

    raise ValueError("Unsupported entity scope.")


def _dependency_blockers(job: DataImportJob, change: DataImportChange) -> list[str]:
    blockers: list[str] = []
    if change.change_type != "create" or not change.rollbackable:
        return blockers
    try:
        target_id = int(change.target_entity_id)
    except (TypeError, ValueError):
        return blockers
    if job.entity_scope == "supplier":
        dependent_items = InventoryItem.query.filter_by(organization_id=job.organization_id, supplier_id=target_id).count()
        dependent_mappings = SupplierItemMapping.query.filter_by(organization_id=job.organization_id, supplier_id=target_id).count()
        if dependent_items or dependent_mappings:
            blockers.append("The supplier now has linked records.")
    elif job.entity_scope in {"inventory_item", "opening_inventory"}:
        invoice_lines = 0
        movements = 0
        counts = 0
        mappings = SupplierItemMapping.query.filter_by(organization_id=job.organization_id, inventory_item_id=target_id).count()
        if invoice_lines or movements or counts or mappings:
            blockers.append("The inventory item now has dependent records.")
    elif job.entity_scope == "supplier_item_mapping":
        invoice_lines = 0
        if invoice_lines:
            blockers.append("The mapping is now referenced by downstream records.")
    return blockers


def _delete_source_record(change: DataImportChange, entity: Any) -> None:
    db.session.delete(entity)


def _rollback_entity(change: DataImportChange, entity: Any) -> None:
    if change.entity_type == "supplier":
        _restore_supplier(dict(change.previous_json or {}), entity)
    elif change.entity_type in {"inventory_item", "opening_inventory"}:
        _restore_inventory_item(dict(change.previous_json or {}), entity)
    elif change.entity_type == "supplier_item_mapping":
        _restore_mapping(dict(change.previous_json or {}), entity)


def _job_response(job: DataImportJob, *, status_code: int = 200):
    return jsonify({"job": _serialize_job_detail(job)}), status_code


@bp.get("/api/imports/organizations/<int:organization_id>/jobs")
@login_required
def list_import_jobs(organization_id: int):
    organization, error = _require_import_access(organization_id)
    if error:
        return error
    jobs = DataImportJob.query.filter_by(organization_id=organization.id).order_by(DataImportJob.created_at.desc(), DataImportJob.id.desc()).all()
    return jsonify({"jobs": [_job_to_summary(job) for job in jobs]}), 200


@bp.post("/api/imports/jobs")
@login_required
def upload_import_job():
    payload_organization_id = request.form.get("organizationId") or (request.get_json(silent=True) or {}).get("organizationId")
    organization_id = _safe_int(payload_organization_id)
    organization, error = _require_import_access(organization_id)
    if error:
        return error

    uploaded = request.files.get("file")
    if uploaded is None:
        return json_error("No import file uploaded.", 400)
    if not uploaded.filename:
        return json_error("The uploaded file is missing a filename.", 400)
    entity_scope = str(request.form.get("entityScope") or (request.get_json(silent=True) or {}).get("entityScope") or "supplier").strip()
    if entity_scope not in SUPPORTED_ENTITY_SCOPES:
        return json_error("Unknown import entity scope.", 400)
    try:
        content = _read_upload_bytes(uploaded)
        columns, rows, source_type = _parse_spreadsheet(uploaded.filename, content)
        source_hash = _sha256(content)
        existing = DataImportJob.query.filter_by(organization_id=organization.id, source_hash=source_hash).first()
        if existing is not None and existing.status != "ROLLED_BACK":
            return json_error("That source file was already uploaded for this organization.", 409)
        if not rows:
            return json_error("The uploaded file does not contain any data rows.", 400)

        job = DataImportJob(
            organization_id=organization.id,
            created_by_user_id=current_user.id,
            source_type=source_type,
            source_file_name=Path(uploaded.filename).name,
            source_file_extension=Path(uploaded.filename).suffix.lower(),
            source_mime_type=uploaded.mimetype or ("text/csv" if source_type == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            source_hash=source_hash,
            storage_path="",
            status="MAPPING_REQUIRED",
            entity_scope=entity_scope,
            mapping_json={"entityScope": entity_scope, "fieldMappings": {}, "fixedValues": {}},
            summary_json={"sourceColumns": columns, "sampleRows": rows[:5]},
            row_count=len(rows),
            batch_id=f"batch-{uuid4().hex[:12]}",
        )
        db.session.add(job)
        db.session.flush()
        source_file = _save_upload(job, uploaded.filename, content)
        job.storage_path = source_file.storage_path
        for index, source_row in enumerate(rows, start=1):
            normalized_row = {str(key): value for key, value in source_row.items()}
            db.session.add(
                DataImportRow(
                    job=job,
                    row_number=index,
                    entity_type=entity_scope,
                    source_row_json=normalized_row,
                    normalized_row_json={},
                    row_fingerprint=_sha256(json.dumps(normalized_row, sort_keys=True, default=str).encode("utf-8")),
                    status="uploaded",
                    target_entity_type=entity_scope,
                    target_entity_id="",
                )
            )
        db.session.commit()
        record_audit_event(
            event_type="import.job_uploaded",
            entity_type="data_import_job",
            entity_id=job.id,
            organization_id=organization.id,
            actor_user_id=current_user.id,
            metadata={"sourceFileName": job.source_file_name, "entityScope": job.entity_scope, "rowCount": job.row_count},
        )
        db.session.commit()
        return _job_response(job, status_code=201)
    except ValueError as exc:
        db.session.rollback()
        return json_error(str(exc), 400)


@bp.get("/api/imports/jobs/<int:job_id>")
@login_required
def get_import_job(job_id: int):
    job, error = _job_or_404(job_id)
    if error:
        return error
    return _job_response(job)


@bp.post("/api/imports/jobs/<int:job_id>/mapping")
@login_required
def set_import_job_mapping(job_id: int):
    job, error = _job_or_404(job_id)
    if error:
        return error
    payload = request.get_json(silent=True) or {}
    try:
        _apply_mapping(job, payload)
        db.session.commit()
        record_audit_event(
            event_type="import.mapping_saved",
            entity_type="data_import_job",
            entity_id=job.id,
            organization_id=job.organization_id,
            actor_user_id=current_user.id,
            metadata={"entityScope": job.entity_scope},
        )
        db.session.commit()
        return _job_response(job)
    except ValueError as exc:
        db.session.rollback()
        return json_error(str(exc), 400)


@bp.post("/api/imports/jobs/<int:job_id>/preview")
@login_required
def preview_import_job(job_id: int):
    job, error = _job_or_404(job_id)
    if error:
        return error
    if not job.mapping_json:
        return json_error("Save a mapping before previewing the import.", 400)
    try:
        _preview_job(job, job.organization)
        db.session.commit()
        record_audit_event(
            event_type="import.preview_generated",
            entity_type="data_import_job",
            entity_id=job.id,
            organization_id=job.organization_id,
            actor_user_id=current_user.id,
            metadata={"blockedRows": job.blocked_row_count, "warningCount": job.warning_count},
        )
        db.session.commit()
        return _job_response(job)
    except ValueError as exc:
        db.session.rollback()
        return json_error(str(exc), 400)


@bp.post("/api/imports/jobs/<int:job_id>/approve")
@login_required
def approve_import_job(job_id: int):
    job, error = _job_or_404(job_id)
    if error:
        return error
    if job.blocked_row_count:
        return json_error("Resolve the validation issues before approving the import.", 400)
    job.status = "APPROVED"
    job.approved_by_user_id = current_user.id
    job.approved_at = utc_now()
    db.session.commit()
    record_audit_event(
        event_type="import.approved",
        entity_type="data_import_job",
        entity_id=job.id,
        organization_id=job.organization_id,
        actor_user_id=current_user.id,
        metadata={"entityScope": job.entity_scope, "rowCount": job.row_count},
    )
    db.session.commit()
    return _job_response(job)


@bp.post("/api/imports/jobs/<int:job_id>/execute")
@login_required
def execute_import_job(job_id: int):
    job, error = _job_or_404(job_id)
    if error:
        return error
    if job.status != "APPROVED":
        return json_error("Approve the import before executing it.", 400)

    try:
        applied_count = 0
        blocked_count = 0
        warning_count = 0
        with db.session.begin_nested():
            _preview_job(job, job.organization)
            if job.blocked_row_count:
                raise ValueError("Resolve the validation issues before executing the import.")
            for row in job.rows:
                if row.blocked_count:
                    blocked_count += 1
                    continue
                change = DataImportChange.query.filter_by(data_import_job_id=job.id, row_fingerprint=row.row_fingerprint).first()
                if change is None:
                    change = DataImportChange(
                        job=job,
                        row=row,
                        entity_type=job.entity_scope,
                        change_type="skip",
                        target_entity_id="",
                        row_fingerprint=row.row_fingerprint,
                        previous_json={},
                        applied_json=dict(row.normalized_row_json or {}),
                        rollbackable=True,
                        status="preview",
                    )
                    db.session.add(change)
                change_type, entity, previous_json, applied_json, rollbackable = _apply_row(job, job.organization, row, current_user.id)
                change.change_type = change_type
                change.target_entity_id = str(getattr(entity, "id", "")) if entity is not None else ""
                change.previous_json = previous_json
                change.applied_json = applied_json
                change.rollbackable = rollbackable
                change.status = "applied"
                row.status = "completed"
                row.issue_summary = "; ".join(issue.message for issue in row.issues)
                applied_count += 1
                warning_count += row.warning_count
            job.applied_row_count = applied_count
            job.warning_count = warning_count
            job.status = "COMPLETED_WITH_WARNINGS" if warning_count or blocked_count else "COMPLETED"
            job.executed_at = utc_now()
            _cleanup_source_file(job)
        db.session.commit()
        record_audit_event(
            event_type="import.executed",
            entity_type="data_import_job",
            entity_id=job.id,
            organization_id=job.organization_id,
            actor_user_id=current_user.id,
            metadata={"entityScope": job.entity_scope, "appliedRows": applied_count, "blockedRows": blocked_count, "warningCount": warning_count},
        )
        db.session.commit()
        return _job_response(job)
    except ValueError as exc:
        db.session.rollback()
        job.status = "FAILED"
        db.session.commit()
        return json_error(str(exc), 400)


@bp.post("/api/imports/jobs/<int:job_id>/rollback")
@login_required
def rollback_import_job(job_id: int):
    job, error = _job_or_404(job_id)
    if error:
        return error
    if job.status not in {"COMPLETED", "COMPLETED_WITH_WARNINGS"}:
        return json_error("Only completed imports can be rolled back.", 400)

    blockers: list[str] = []
    applied_changes = DataImportChange.query.filter_by(data_import_job_id=job.id, status="applied").order_by(DataImportChange.id.desc()).all()
    for change in applied_changes:
        blockers.extend(_dependency_blockers(job, change))
    if blockers:
        job.rollback_blockers_json = blockers
        db.session.commit()
        return json_error("This import cannot be rolled back safely yet.", 409, errors={"blockers": "; ".join(blockers)})

    try:
        for change in applied_changes:
            if change.change_type == "skip" or not change.target_entity_id:
                continue
            if job.entity_scope == "supplier":
                entity = Supplier.query.filter_by(id=int(change.target_entity_id), organization_id=job.organization_id).first()
            elif job.entity_scope in {"inventory_item", "opening_inventory"}:
                entity = InventoryItem.query.filter_by(id=int(change.target_entity_id), organization_id=job.organization_id).first()
            else:
                entity = SupplierItemMapping.query.filter_by(id=int(change.target_entity_id), organization_id=job.organization_id).first()
            if entity is None:
                continue
            if change.change_type == "create":
                _delete_source_record(change, entity)
            else:
                _rollback_entity(change, entity)
            change.status = "rolled_back"
        job.status = "ROLLED_BACK"
        job.rolled_back_at = utc_now()
        _cleanup_source_file(job)
        db.session.commit()
        record_audit_event(
            event_type="import.rolled_back",
            entity_type="data_import_job",
            entity_id=job.id,
            organization_id=job.organization_id,
            actor_user_id=current_user.id,
            metadata={"entityScope": job.entity_scope},
        )
        db.session.commit()
        return _job_response(job)
    except ValueError as exc:
        db.session.rollback()
        return json_error(str(exc), 400)
